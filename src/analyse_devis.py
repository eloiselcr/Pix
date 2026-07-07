import json
import openpyxl # pour manipuler excel
from pathlib import Path # pour viser /Pix
from pypdf import PdfReader # pour lire le pdf et extract
from google import genai # pour IA


# ======== Init des chemins + Clé API =============

# On retrouve nos chemins de fichiers
chemin_racine = Path(__file__).parent.parent # vise /Pix
chemin_config = chemin_racine / "config.json"
chemin_pdf = chemin_racine / "devis_test.pdf"


# test_ai
# On lit notre clé API
with open(chemin_config, "r", encoding="utf-8") as fichier:
    config = json.load(fichier)
cle_api = config.get("GEMINI_API_KEY")
if not cle_api:
    print("Erreur : Aucune clé API trouvée dans config.json")
    exit()


# ======== Extract du PDF =============


def extraire_texte_pdf(chemin_du_fichier):
    if not chemin_du_fichier.exists():
        print(f"Erreur : Le fichier {chemin_du_fichier.name} n'existe pas à la racine du projet !")
        exit()
        
    print(f"Ouverture du fichier : {chemin_du_fichier.name}")
    
    # classe PdfReader -> lit devis_pdf
    devis_pdf = PdfReader(chemin_du_fichier)
    
    # string pdf extract pour stocker texte extrait
    pdf_extract = ""
    
    # boucle pour parcourir devis_pdf et extraire
    for index, page in enumerate(devis_pdf.pages): # on parcourt toutes les pages
        print(f"Lecture de la page {index + 1}...")
        texte_page = page.extract_text() # fonction extract 
        if texte_page: # si la page n'est pas vide
            pdf_extract = pdf_extract + texte_page + "\n" # on ajoute les nouvelles pages au fur et à mesure
            
    print("\n--- Texte extrait du PDF : ---")
    print(pdf_extract)
    print("--- Fin du texte extrait ---")
    return pdf_extract


texte_devis = extraire_texte_pdf(chemin_pdf)


# ======== Prompt a l'IA =============


client = genai.Client(api_key=cle_api)

consigne = """
Tu es un assistant comptable expert en extraction de données financières. Analyse le texte brut d'un devis pour en extraire des informations structurées au format JSON.

### RÈGLES D'EXTRACTION CRUCIALES :
1. **Numéro du devis** : Extrais le numéro de référence du devis (ex: C2098/SD/0525).
2. **Résumé du devis** : Fais un résumé très court en une phrase simple de l'objectif principal du devis. Ce résumé ne doit pas être mot pour mot ce qui est indiqué dans les unités d'oeuvres.
3. **Unités d'Œuvre (UO)** : Repère les lignes de prestations (souvent présentées en tableau). Chaque ligne doit être traitée indépendamment sans mélanger les informations. Si le devis ne contient qu’une seule prestation, il se peut que les informations ne soient pas présentées en tableau mais dans une ligne. C’est à toi de faire attention aux détails.

Les UO sont souvent composées ainsi : 
- **Nom de l'UO (ref_bpu)** : Utilise la catégorisation stricte (ex: EVOL M-MC, EVOL P-C). Si un titre de la prestation est présent (ex: AXO-XXX - Nom ; ou du texte sur la ligne), ajoute le.
- **Prix HT** : Extrais le montant sous forme de nombre décimal pur (sans le symbole €).
- **Quantité** : Cherche UNIQUEMENT si une colonne ou une mention explicite le mot "Quantité". Si le mot exact "Quantité" n'est pas spécifié pour la ligne, attribue la valeur par défaut : 1.
   - Ignore les mentions alternatives comme "Nombre d’écrans", "Volume d'objets", etc.
- **Date de livraison / Validité** : Extrais la date limite de validité de la proposition (ex: "valide jusqu’au 31 juillet 2026 inclus" -> extraire la date). Tu DOIS chercher cette date (qu’elle soit écrite en forme numérique ou en toutes lettres). Généralement  Convertis-la obligatoirement au format JJ-MM-AAAA. Si aucune date n'est trouvable, renvoie null.

### FORMAT DE SORTIE STRICT :
Renvoie UNIQUEMENT le bloc JSON ci-dessous, sans texte d'introduction, sans explication, et sans balises de code Markdown.

{
  "numero_devis": "chaîne de caractères ou null",
  "resume_devis": "chaîne de caractères",
  "unites_oeuvre": [
    {
      "ref_bpu": "chaîne de caractères",
      "prix_HT": 0.0,
      "quantite": 1,
      "date_livraison": "JJ-MM-AAAA ou null"
    }
  ]
}

### TEXTE DU DEVIS À ANALYSER :
Voici le texte que tu as à analyser : 

"""

prompt = consigne + texte_devis

print("Envoi du devis à l'IA...")

reponse = client.models.generate_content(
    model='gemini-2.5-flash',
    contents=prompt,
    config=genai.types.GenerateContentConfig(
        response_mime_type="application/json",
    ),
)

print("---------------------------------")
print("Réponse de l'IA : ")
print(reponse.text)


# JSON to Dictionnaire Python
donnees_devis = json.loads(reponse.text)

print("\nType de la variable donnees_devis : ", type(donnees_devis))

print("Référence du devis extraite : ", donnees_devis["numero_devis"])
print("Résumé rapide du devis : ", donnees_devis["resume_devis"])


# ======== MAPPING EXCEL =============

# on peut pas juste faire JSON -> excel

# création d'un dictionnaire de correspondance donnees_excel
# clé (balise excel) : valeur (balise JSON IA)
donnees_excel = {
    "observations": f"Devis {donnees_devis['numero_devis']} - {donnees_devis['resume_devis']}"
}


liste_uo = donnees_devis["unites_oeuvre"]

# parcourir les UO, max UO FEB = 8 
for i in range (1, 8):
    if i <= len(liste_uo):
        uo = liste_uo[i - 1]

        donnees_excel[f"ref_bpu_ligne{i}"] = f"Devis {donnees_devis['numero_devis']} - {uo['ref_bpu']}"
        donnees_excel[f"prix_HT_ligne{i}"] = uo["prix_HT"]
        donnees_excel[f"quantite_ligne{i}"] = uo["quantite"]
        donnees_excel[f"date_livraison_ligne{i}"] = uo["date_livraison"]

    else: 
        # si c'est vide, on met rien
        donnees_excel[f"ref_bpu_ligne{i}"] = ""
        donnees_excel[f"prix_HT_ligne{i}"] = ""
        donnees_excel[f"quantite_ligne{i}"] = ""
        donnees_excel[f"date_livraison_ligne{i}"] = ""
        
        
print("\n--- Vérification du dictionnaire final : ---")
print(donnees_excel)



# ============= INSERTION EXCEL ==================

chemin_modele = chemin_racine / "templates" / "FEB_modele.xlsx"
chemin_sortie = chemin_racine / "out" / "FEB_generee.xlsx"
chemin_sortie.parent.mkdir(parents=True, exist_ok=True)


print(f"Ouvertue du modèle excel : {chemin_modele.name}...")

# load_workbook -> ouvre le modèle en mémoire (modèle = "workbook" en openpyxl)
wb = openpyxl.load_workbook(chemin_modele)
# indique la feuille de travail
onglet = wb["FEB"]

print("Remplacement des balises...")

# quadrillage de la feuille
for ligne in onglet.iter_rows(): # pour chaque ligne, du haut vers bas
    for cellule in ligne: # on regarde chaque cellule d'une ligne (A1, A2, A3...)
        valeur = cellule.value # on récupère la valeur dans la celulle

        if isinstance(valeur, str): # si la balise est remplie (str présent)
            for cle, val_remplacement in donnees_excel.items(): # on parcourt notre dico donnees_excel
                balise = f"{{{{{cle}}}}}"

                # Cas 1 : la balise est identique à la valeur 
                # ex : pour prix_HT_ligne1, on ne mets qu'une valeur
                if valeur == balise:
                    cellule.value = val_remplacement
                    break
                # Cas 2 : la balise se trouve avec du texte
                # ex : N° {{num_feb}} -> N° est du texte
                elif balise in valeur: # valeur = du texte. La base se trouve avec du texte
                    valeur = valeur.replace(balise, str(val_remplacement))
                    cellule.value = valeur


print(f"Fichier généré trouvable sous : {chemin_sortie.name} !")
wb.save(chemin_sortie)

print("Terminé !")