import openpyxl
from pathlib import Path

# 1. On définit le chemin vers ton modèle
chemin_racine = Path(__file__).parent.parent
chemin_modele = chemin_racine / "templates" / "FEB_modele.xlsx"

print(f"Chargement du fichier : {chemin_modele.name}")

# 2. openpyxl ouvre le fichier Excel en mémoire
wb = openpyxl.load_workbook(chemin_modele)

# 3. On va parcourir tous les onglets du fichier Excel
for nom_onglet in wb.sheetnames:
    onglet = wb[nom_onglet]
    print(f"\nRecherche dans l'onglet : {nom_onglet}")
    
    # 4. On parcourt le tableau cellule par cellule
    for ligne in onglet.iter_rows():
        for cellule in ligne:
            valeur = cellule.value
            
            # 5. Si la cellule contient du texte avec "{{"
            if isinstance(valeur, str) and "{{" in valeur:
                print(f"  Cellule {cellule.coordinate} : {valeur}")
